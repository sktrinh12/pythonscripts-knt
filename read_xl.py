from openpyxl import load_workbook
import re
import collections
import json
import pathlib

dir_path = pathlib.Path(
    "/Users/spencer.trinhkinnate.com/OneDrive - Kinnate Biopharma Inc/Attachments/"
)
extension = "*.xlsx"
pattern = r"^FT\d+-\d+\S+"

xlsx_files = list(dir_path.glob(extension))

# print(xlsx_files)


def parse_xlsx(file_name):
    workbook = load_workbook(
        filename=file_name,
        read_only=True,
    )

    wshs = [sh for sh in workbook.sheetnames]
    cell_vals = []
    for wsh in wshs:
        wbk = workbook[wsh]
        for row in wbk.iter_rows():
            for cell in row:
                match = re.match(pattern, str(cell.value))
                if match:
                    cell_vals.append(match.group(0))

    counter = collections.Counter(cell_vals)
    count_dict = counter.items()
    count_vals = list(count_dict)
    # max_count = counter.most_common(1)[0][1]
    # max_keys = [k for k, v in count_dict if v == max_count]
    print(json.dumps(count_vals, indent=4))
    # print(max_keys[0])


if __name__ == "__main__":
    for file_name in xlsx_files:
        print(file_name)
        parse_xlsx(file_name)
        print("-" * 20)
    # print(xlsx_files[7])
    # parse_xlsx(xlsx_files[7])
